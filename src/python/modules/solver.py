import signal
import os
import clingo
import re
from threading import Timer
from collections import defaultdict
from openpyxl import Workbook
from rich.console import Console
console = Console()

from modules.gestori import (
    GestoreMappe
)

# Percorso principale del progetto
base_dir = os.path.abspath(os.path.join(
    os.path.dirname(__file__), "../../../src"))

# Directory per gli script ASP
asp_dir = os.path.join(base_dir, "asp")
facts_dir = os.path.join(asp_dir, "facts")
results_dir = os.path.join(asp_dir, "results")

# File principali
files_to_include = [
    os.path.join(asp_dir, "main.lp"),
    os.path.join(asp_dir, "preferenze.lp")
]

# Funzione per leggere i file dalla directory dei facts


def read_files(directory):
    content = ""
    console.print(f"Lettura file dalla directory: {directory}")
    for filename in sorted(os.listdir(directory)):  # Ordine deterministico
        if filename.endswith(".asp"):
            console.print(f"Leggendo il file: {filename}")
            with open(os.path.join(directory, filename), "r") as file:
                content += file.read() + "\n"
    return content

# Leggi il contenuto dei file ASP


def load_program():
    # Leggi i facts
    facts_content = read_files(facts_dir)
    # Leggi i file principali
    program_content = ""
    for file in files_to_include:
        if not os.path.exists(file):
            raise FileNotFoundError(f"File non trovato: {file}")
        with open(file, "r") as f:
            program_content += f.read() + "\n"
    return program_content + facts_content


# Flag per interrompere il solving
stop_solving = False

def signal_handler(signum, frame):
    """Gestore dei segnali per Ctrl+C."""
    global stop_solving
    stop_solving = True
    console.print(
        "[bold yellow]Ricevuto Ctrl+C: interruzione in corso...[/bold yellow]")


def timeout_handler(ctl):
    console.print(
        "[bold red]Timeout raggiunto. Interruzione in corso...[/bold red]")
    ctl.interrupt()


def solve_program(mode="full", verbose=True, arguments=[]):
    """
    Esegue il solver con la modalità specificata.

    Args:
        mode (str): Modalità di esecuzione. Può essere "full", "gringo", o "clingo".
        verbose (bool): Se True, stampa i dettagli delle soluzioni.
        arguments (list): Argomenti da passare al controllo di Clingo.
    """
    global stop_solving
    stop_solving = False  # Resetta il flag all'inizio
    # Imposta il gestore per Ctrl+C
    signal.signal(signal.SIGINT, signal_handler)

    try:
        complete_program = load_program()  # Funzione che carica il programma ASP
        results_file = os.path.join(results_dir, "solution.txt")

        # Pulisce il file dei risultati se esiste
        if os.path.exists(results_file):
            os.remove(results_file)

        def on_model(model):
            if stop_solving:
                ctl.interrupt()  # Interrompe la ricerca in corso
                raise KeyboardInterrupt  # Genera un'interruzione

            # model_symbols = model.symbols(atoms=True)
            model_symbols = model.symbols(shown=True)
            model_costs = model.cost
            with open(results_file, "w") as f:
                f.write(f"model: {model_symbols}\n")
                if model_costs:
                    f.write(f"cost: {model_costs}\n")
            if verbose:
                console.print(
                    f"[bold green]Model trovato: [/bold green]")
                print(model_symbols)

        ctl = clingo.Control(arguments)
        ctl.add("base", [], complete_program)
        ctl.ground([("base", [])])

        console.print("[bold blue]Avvio del solving...[/bold blue]")

        # Imposta il timer per il timeout
        timer = Timer(5, timeout_handler, args=(ctl,))
        timer.start()

        result = ctl.solve(on_model=on_model)

        # Ferma il timer se la soluzione termina prima del timeout
        timer.cancel()

        if stop_solving:
            console.print(
                "[bold yellow]Esecuzione interrotta dall'utente.[/bold yellow]")
        elif result.satisfiable:
            console.print(
                f"[bold green]Soluzione trovata![/bold green] Risultati salvati in: [magenta]{results_file}[/magenta]")
        else:
            console.print("[bold red]Nessuna soluzione trovata.[/bold red]")

    except KeyboardInterrupt:
        console.print(
            "[bold yellow]Esecuzione interrotta manualmente.[/bold yellow]")
    except FileNotFoundError as e:
        console.print(f"[bold red]Errore:[/bold red] {e}")
    except Exception as e:
        console.print(f"[bold red]Errore durante il solving:[/bold red] {e}")
    finally:
        # Ripristina il comportamento predefinito
        signal.signal(signal.SIGINT, signal.SIG_DFL)

# Funzione per estrarre i docenti e i corsi dal modello


def extract_data(model):
    pattern = r"garante\(docente\((\d+)\),corso\((\d+)\)\)"
    matches = re.findall(pattern, model)
    return matches


def write_table(input_file, output_file):

    input_file = os.path.join(results_dir, input_file)
    output_file = os.path.join(results_dir, output_file)

    with open(input_file, "r") as file:
        # Legge la prima riga e rimuove eventuali spazi bianchi
        model_symbols = file.readline().strip()

    # Estrai i dati dal modello
    extracted_data = extract_data(model_symbols)

    # Raggruppa i docenti per corso
    grouped_by_course = defaultdict(list)
    for docente, corso in extracted_data:
        grouped_by_course[corso].append(docente)

    # Ordina i dati per codice corso
    extracted_data = sorted(extracted_data, key=lambda x: int(x[1]))

    # Creazione del file Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Garanti per Corso"

    # Scrittura dell'intestazione della tabella
    headers = ["Docente", "Codice Docente",
               "Codice Corso", "Nome Corso", "SSD 2015"]
    for col_index, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_index, value=header)

    # GestoreMappe.inizializza()
    mappa_docenti = GestoreMappe.get_mappa_docenti()
    mappa_docenti_settori = GestoreMappe.get_mappa_docenti_settori()
    mappa_ssd = GestoreMappe.get_mappa_ssd_2024_2015()
    mappa_corsi_nomi = GestoreMappe.get_mappa_corsi_nomi()

    # Scrittura dei dati nella tabella
    row_index = 2
    for docente, corso in extracted_data:
        docente = int(docente)
        corso = int(corso)

        # Ottieni le informazioni dal dizionario
        docente_nome = mappa_docenti[docente]
        settore = mappa_docenti_settori[docente]
        if settore != 'null':
            settore = mappa_ssd[settore]
        nome_corso = mappa_corsi_nomi[corso]

        # Scrivi i dati in una riga
        # Nome del docente
        ws.cell(row=row_index, column=1, value=docente_nome)
        ws.cell(row=row_index, column=2, value=docente)       # Codice docente
        ws.cell(row=row_index, column=3, value=corso)         # Codice corso
        ws.cell(row=row_index, column=4, value=nome_corso)       # Nome Corso
        ws.cell(row=row_index, column=5, value=settore)       # SSD 2024
        row_index += 1

    # Salva il file Excel
    wb.save(output_file)
    console.print(
        f"[bold green]Tabella generata con successo![/bold green] Risultati salvati in: [magenta]{output_file}[/magenta]")
